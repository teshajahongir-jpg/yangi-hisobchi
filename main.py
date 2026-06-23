import logging
import sqlite3
import os
import random
import asyncio
from datetime import datetime, date, timedelta
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes, ConversationHandler, CallbackQueryHandler
)
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from aiohttp import web
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from namoz_data import bugungi_namoz_vaqtlari

# ─── SOZLAMALAR ───────────────────────────────────────────────
BOT_TOKEN = "8946241677:AAG0wsbr_83HZTeXqw_HBkTnzrgD_RsRCOc"
ADMIN_ID = 8252424738

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "xarajat.db")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Global scheduler va application
scheduler = None
application = None

KIRIM_MIQDOR, KIRIM_TAVSIF = range(2)
CHIQIM_MIQDOR, CHIQIM_KATEGORIYA = range(2, 4)
LIMIT_KIRITISH = 4
BYUDJET_KIRITISH = 5
QARZ_ISM, QARZ_MIQDOR, QARZ_TUR, QARZ_ESLATMA, QARZ_IZOH = range(10, 15)
SANA_ISM, SANA_SANA, SANA_TAVSIF = range(15, 18)
XABAR_MATN = 18

KATEGORIYALAR = [
    "🍔 Oziq-ovqat", "🚌 Transport", "🏠 Uy-joy", "👕 Kiyim",
    "💊 Sog'liq", "📚 Ta'lim", "🎉 Dam olish", "📦 Boshqa"
]

TAYYOR_MIQDORLAR = [
    "5,000", "10,000", "20,000", "50,000",
    "100,000", "200,000", "500,000", "✍️ O'zim yozaman"
]

MOTIVATSION_XABARLAR = [
    "☀️ Assalomu alaykum! Bugun ham samarali kun bo'lsin!",
    "🌟 Yangi kun — yangi imkoniyat! Maqsadlaringizga qadam tashlang!",
    "💪 Bugun kechagidan yaxshiroq bo'lish uchun imkoniyatingiz bor!",
    "🤲 Alloh barchangizga barakali kun bersin!",
    "✨ Bugun kichik qadamlar katta natijalarga olib boradi!",
    "🌱 Har bir yaxshi ish bilan kun boshlansin!",
    "💡 Bugun bitta yangi narsa o'rganing!",
    "🎯 Maqsadingizga har kuni bir qadam yaqinlashing!",
    "🌈 Sabr qiling — yaxshi kunlar albatta keladi!",
    "🔥 Bugun o'z rekordingizni yangilang!",
]

SOGLIK_MASLAHATLARI = [
    "💧 Bugun kamida 8 stakan suv iching!",
    "🚶 30 daqiqa piyoda yuring — yurak uchun juda foydali!",
    "🥗 Bugun ovqatga ko'proq sabzavot qo'shing!",
    "😴 Sifatli uxlash — eng yaxshi dori. Erta yoting!",
    "🧘 5 daqiqa chuqur nafas oling — stress kamayadida!",
    "🍎 Kuniga bir meva yeng — tabib kerak bo'lmaydi!",
    "📵 Uxlashdan 1 soat oldin telefon ekranini o'chiring!",
    "🦷 Tishlaringizni kuniga 2 marta tozalang!",
    "☀️ Kuniga 15-20 daqiqa quyosh nurida bo'ling!",
    "🏃 Muntazam jismoniy faollik — uzoq umrning sirri!",
]


# ─── DATABASE ─────────────────────────────────────────────────
def db_connect():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = db_connect()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            ism TEXT, username TEXT,
            royxatdan_otgan TEXT,
            kunlik_limit REAL DEFAULT 0,
            oylik_byudjet REAL DEFAULT 0,
            eslatmalar INTEGER DEFAULT 1
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, tur TEXT, miqdor REAL,
            kategoriya TEXT, tavsif TEXT, sana TEXT, vaqt TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS limit_ogohlantirish (
            user_id INTEGER, sana TEXT, yuborilgan INTEGER DEFAULT 0,
            PRIMARY KEY (user_id, sana)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS qarzlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, ism TEXT, miqdor REAL, tur TEXT,
            sana TEXT, qaytarish_sanasi TEXT, izoh TEXT,
            qaytarildi INTEGER DEFAULT 0
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS muhim_sanalar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, ism TEXT, sana TEXT, tavsif TEXT
        )
    """)
    conn.commit()
    conn.close()


def royxatdan_otkazish(user_id, ism, username):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,))
    if not c.fetchone():
        c.execute(
            "INSERT INTO users (user_id, ism, username, royxatdan_otgan) VALUES (?,?,?,?)",
            (user_id, ism or "", username or "", str(date.today()))
        )
        conn.commit()
    conn.close()


def barcha_userlar():
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT user_id, ism, username, kunlik_limit, eslatmalar FROM users")
    rows = c.fetchall()
    conn.close()
    return rows


def qosh_tranzaksiya(user_id, tur, miqdor, kategoriya, tavsif):
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        "INSERT INTO transactions (user_id, tur, miqdor, kategoriya, tavsif, sana, vaqt) VALUES (?,?,?,?,?,?,?)",
        (user_id, tur, miqdor, kategoriya, tavsif, str(date.today()), datetime.now().strftime("%H:%M"))
    )
    conn.commit()
    conn.close()


def oxirgi_yozuvni_ochirish(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT id, tur, miqdor, tavsif FROM transactions WHERE user_id=? ORDER BY id DESC LIMIT 1", (user_id,))
    row = c.fetchone()
    if row:
        c.execute("DELETE FROM transactions WHERE id=?", (row[0],))
        conn.commit()
    conn.close()
    return row


def bugungi_hisobot_db(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        "SELECT tur, miqdor, kategoriya, tavsif, vaqt FROM transactions WHERE user_id=? AND sana=? ORDER BY id",
        (user_id, str(date.today()))
    )
    rows = c.fetchall()
    conn.close()
    return rows


def bugungi_chiqim_jami(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        "SELECT COALESCE(SUM(miqdor),0) FROM transactions WHERE user_id=? AND sana=? AND tur='chiqim'",
        (user_id, str(date.today()))
    )
    jami = c.fetchone()[0]
    conn.close()
    return jami


def oylik_hisobot_db(user_id, oy=None):
    conn = db_connect()
    c = conn.cursor()
    if oy is None:
        oy = datetime.now().strftime("%Y-%m")
    c.execute(
        "SELECT tur, SUM(miqdor) FROM transactions WHERE user_id=? AND sana LIKE ? GROUP BY tur",
        (user_id, f"{oy}%")
    )
    rows = c.fetchall()
    conn.close()
    return rows


def davr_hisobot_db(user_id, boshlanish):
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        "SELECT tur, miqdor, kategoriya, tavsif, sana, vaqt FROM transactions "
        "WHERE user_id=? AND sana>=? ORDER BY sana, id",
        (user_id, str(boshlanish))
    )
    rows = c.fetchall()
    conn.close()
    return rows


def kategoriya_hisobot_db(user_id):
    conn = db_connect()
    c = conn.cursor()
    oy = datetime.now().strftime("%Y-%m")
    c.execute(
        "SELECT kategoriya, SUM(miqdor) FROM transactions "
        "WHERE user_id=? AND tur='chiqim' AND sana LIKE ? GROUP BY kategoriya ORDER BY SUM(miqdor) DESC",
        (user_id, f"{oy}%")
    )
    rows = c.fetchall()
    conn.close()
    return rows


def limit_olish(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT kunlik_limit FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else 0


def limit_saqlash(user_id, limit):
    conn = db_connect()
    c = conn.cursor()
    c.execute("UPDATE users SET kunlik_limit=? WHERE user_id=?", (limit, user_id))
    conn.commit()
    conn.close()


def byudjet_olish(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT oylik_byudjet FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else 0


def byudjet_saqlash(user_id, byudjet):
    conn = db_connect()
    c = conn.cursor()
    c.execute("UPDATE users SET oylik_byudjet=? WHERE user_id=?", (byudjet, user_id))
    conn.commit()
    conn.close()


def eslatma_holati(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT eslatmalar FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else 1


def eslatma_almashtirish(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT eslatmalar FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    yangi = 0 if (row and row[0] == 1) else 1
    c.execute("UPDATE users SET eslatmalar=? WHERE user_id=?", (yangi, user_id))
    conn.commit()
    conn.close()
    return yangi


def limit_ogohlantirish_yuborilganmi(user_id, sana):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT yuborilgan FROM limit_ogohlantirish WHERE user_id=? AND sana=?", (user_id, str(sana)))
    row = c.fetchone()
    conn.close()
    return bool(row and row[0] == 1)


def limit_ogohlantirish_belgilash(user_id, sana):
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        "INSERT INTO limit_ogohlantirish (user_id, sana, yuborilgan) VALUES (?,?,1) "
        "ON CONFLICT(user_id, sana) DO UPDATE SET yuborilgan=1",
        (user_id, str(sana))
    )
    conn.commit()
    conn.close()


def admin_statistika():
    conn = db_connect()
    c = conn.cursor()
    bugun = str(date.today())
    oy = datetime.now().strftime("%Y-%m")
    c.execute("SELECT COUNT(*) FROM users")
    jami = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT user_id) FROM transactions WHERE sana=?", (bugun,))
    faol = c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(miqdor),0) FROM transactions WHERE tur='kirim' AND sana LIKE ?", (f"{oy}%",))
    kirim = c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(miqdor),0) FROM transactions WHERE tur='chiqim' AND sana LIKE ?", (f"{oy}%",))
    chiqim = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM qarzlar WHERE qaytarildi=0")
    qarzlar = c.fetchone()[0]
    conn.close()
    return jami, faol, kirim, chiqim, qarzlar


def qarz_qosh_db(user_id, ism, miqdor, tur, qaytarish, izoh):
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        "INSERT INTO qarzlar (user_id, ism, miqdor, tur, sana, qaytarish_sanasi, izoh, qaytarildi) VALUES (?,?,?,?,?,?,?,0)",
        (user_id, ism, miqdor, tur, str(date.today()), qaytarish, izoh)
    )
    conn.commit()
    conn.close()


def qarz_royxat_db(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        "SELECT id, ism, miqdor, tur, qaytarish_sanasi, izoh FROM qarzlar "
        "WHERE user_id=? AND qaytarildi=0 ORDER BY qaytarish_sanasi",
        (user_id,)
    )
    rows = c.fetchall()
    conn.close()
    return rows


def qarz_yop(qarz_id, user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT ism, miqdor, tur FROM qarzlar WHERE id=? AND user_id=?", (qarz_id, user_id))
    row = c.fetchone()
    if row:
        c.execute("UPDATE qarzlar SET qaytarildi=1 WHERE id=? AND user_id=?", (qarz_id, user_id))
        conn.commit()
    conn.close()
    return row


def muddati_yaqin_qarzlar():
    conn = db_connect()
    c = conn.cursor()
    uch_kun = str(date.today() + timedelta(days=3))
    bugun = str(date.today())
    c.execute(
        "SELECT user_id, id, ism, miqdor, tur, qaytarish_sanasi FROM qarzlar "
        "WHERE qaytarildi=0 AND qaytarish_sanasi<=? AND qaytarish_sanasi>=?",
        (uch_kun, bugun)
    )
    rows = c.fetchall()
    conn.close()
    return rows


def sana_qosh_db(user_id, ism, sana, tavsif):
    conn = db_connect()
    c = conn.cursor()
    c.execute("INSERT INTO muhim_sanalar (user_id, ism, sana, tavsif) VALUES (?,?,?,?)",
              (user_id, ism, sana, tavsif))
    conn.commit()
    conn.close()


def sanalar_royxat_db(user_id):
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT id, ism, sana, tavsif FROM muhim_sanalar WHERE user_id=? ORDER BY sana", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows


def ertangi_sanalar():
    conn = db_connect()
    c = conn.cursor()
    ertaga = (date.today() + timedelta(days=1)).strftime("%m-%d")
    c.execute(
        "SELECT user_id, ism, sana, tavsif FROM muhim_sanalar WHERE strftime('%m-%d', sana)=?",
        (ertaga,)
    )
    rows = c.fetchall()
    conn.close()
    return rows


# ─── NAMOZ VAQTLARI ───────────────────────────────────────────
def namoz_vaqtlarini_ol():
    bugun = str(date.today())
    vaqtlar = bugungi_namoz_vaqtlari(bugun)
    if vaqtlar:
        return vaqtlar
    try:
        d = datetime.now()
        url = f"https://api.aladhan.com/v1/timings/{d.day}-{d.month}-{d.year}"
        params = {"latitude": 39.7747, "longitude": 64.4286, "method": 8}
        resp = requests.get(url, params=params, timeout=10)
        t = resp.json()["data"]["timings"]
        return {
            "bomdod": t["Fajr"][:5], "peshin": t["Dhuhr"][:5],
            "asr": t["Asr"][:5], "shom": t["Maghrib"][:5], "xufton": t["Isha"][:5],
        }
    except Exception as e:
        logger.error(f"Namoz API xatosi: {e}")
        return None


def keyingi_namoz():
    vaqtlar = namoz_vaqtlarini_ol()
    if not vaqtlar:
        return "—", "—"
    hozir = datetime.now().strftime("%H:%M")
    emoji = {"bomdod": "🌙", "peshin": "☀️", "asr": "🌤", "shom": "🌇", "xufton": "🌃"}
    for nom, vaqt in vaqtlar.items():
        if vaqt > hozir:
            return f"{emoji.get(nom, '')} {nom.capitalize()}", vaqt
    return "🌙 Bomdod (ertaga)", vaqtlar.get("bomdod", "—")


def vaqt_qosh(soat, daqiqa, delta):
    umumiy = (soat * 60 + daqiqa + delta) % (24 * 60)
    return umumiy // 60, umumiy % 60


# ─── ESLATMALAR ───────────────────────────────────────────────
async def barcha_userlarga_yuborish(app, matn):
    for user_id, ism, username, limit_q, eslatma in barcha_userlar():
        if eslatma != 1:
            continue
        try:
            await app.bot.send_message(chat_id=user_id, text=matn)
        except Exception as e:
            logger.warning(f"{user_id}: {e}")


async def namoz_eslatma(app, namoz_nomi):
    await barcha_userlarga_yuborish(app, f"🕌 {namoz_nomi} namoziga 15 daqiqa qoldi! Alloh qabul qilsin 🤲")


async def namoz_oqildimi_sorov(app, namoz_nomi):
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Ha, o'qidim", callback_data=f"namoz_ha_{namoz_nomi}"),
        InlineKeyboardButton("❌ Yo'q", callback_data=f"namoz_yoq_{namoz_nomi}")
    ]])
    for user_id, ism, username, limit_q, eslatma in barcha_userlar():
        if eslatma != 1:
            continue
        try:
            await app.bot.send_message(
                chat_id=user_id,
                text=f"🕌 {namoz_nomi} namozini o'qib bo'ldingizmi?",
                reply_markup=keyboard
            )
        except Exception as e:
            logger.warning(f"{user_id}: {e}")


async def uyqu_eslatma(app):
    await barcha_userlarga_yuborish(app, "😴 Uxlash vaqti keldi! Yaxshi tun 🌙")


async def uyg_onish_eslatma(app):
    await barcha_userlarga_yuborish(app, "🌅 Bomdod vaqti yaqinlashdi! Uyg'oning! 🤲")


async def motivatsion_xabar(app):
    await barcha_userlarga_yuborish(app, random.choice(MOTIVATSION_XABARLAR))


async def soglik_maslahati(app):
    await barcha_userlarga_yuborish(app, "💊 Bugungi sog'liq maslahati:\n\n" + random.choice(SOGLIK_MASLAHATLARI))


async def tungi_hisobot_yuborish(app):
    for user_id, ism, username, limit_q, eslatma in barcha_userlar():
        rows = bugungi_hisobot_db(user_id)
        if not rows:
            continue
        kirim = sum(r[1] for r in rows if r[0] == "kirim")
        chiqim = sum(r[1] for r in rows if r[0] == "chiqim")
        matn = f"📊 Kunlik hisobot — {date.today().strftime('%d.%m.%Y')}\n"
        matn += "─────────────────\n"
        matn += f"✅ Kirim:  {kirim:,.0f} som\n"
        matn += f"❌ Chiqim: {chiqim:,.0f} som\n"
        matn += "─────────────────\n"
        matn += f"💰 Qoldiq: {kirim-chiqim:,.0f} som\n"
        try:
            await app.bot.send_message(chat_id=user_id, text=matn)
        except Exception as e:
            logger.warning(f"{user_id}: {e}")


async def qarz_eslatma_yuborish(app):
    qarzlar = muddati_yaqin_qarzlar()
    user_map = {}
    for uid, qid, ism, miqdor, tur, sana in qarzlar:
        user_map.setdefault(uid, []).append((qid, ism, miqdor, tur, sana))
    for uid, lst in user_map.items():
        matn = "⏰ Qarz eslatmasi!\n\n"
        for qid, ism, miqdor, tur, sana in lst:
            matn += f"{'📤' if tur == 'berdi' else '📥'} {ism} — {miqdor:,.0f} som | {sana}\n"
        try:
            await app.bot.send_message(chat_id=uid, text=matn)
        except Exception as e:
            logger.warning(f"{uid}: {e}")


async def muhim_sana_eslatma(app):
    for uid, ism, sana, tavsif in ertangi_sanalar():
        matn = f"🎂 Ertaga muhim sana!\n{ism} — {sana}"
        if tavsif:
            matn += f"\n{tavsif}"
        try:
            await app.bot.send_message(chat_id=uid, text=matn)
        except Exception as e:
            logger.warning(f"{uid}: {e}")


# ─── REJALASHTIRISH ───────────────────────────────────────────
def rejalashtirish(sched, app):
    sched.add_job(uyqu_eslatma, "cron", hour=22, minute=30, args=[app], id="uyqu", replace_existing=True)
    sched.add_job(tungi_hisobot_yuborish, "cron", hour=0, minute=0, args=[app], id="hisobot", replace_existing=True)
    sched.add_job(motivatsion_xabar, "cron", hour=7, minute=0, args=[app], id="motivatsiya", replace_existing=True)
    sched.add_job(soglik_maslahati, "cron", hour=12, minute=0, args=[app], id="soglik", replace_existing=True)
    sched.add_job(qarz_eslatma_yuborish, "cron", hour=9, minute=0, args=[app], id="qarz_eslatma", replace_existing=True)
    sched.add_job(muhim_sana_eslatma, "cron", hour=8, minute=0, args=[app], id="sana_eslatma", replace_existing=True)
    sched.add_job(namoz_rejalashtir, "cron", hour=3, minute=30, args=[sched, app], id="namoz_yangi", replace_existing=True)


def namoz_rejalashtir(sched, app):
    vaqtlar = namoz_vaqtlarini_ol()
    if not vaqtlar:
        return
    for namoz, vaqt_str in vaqtlar.items():
        try:
            soat, daqiqa = map(int, vaqt_str.split(":"))
            oldin_s, oldin_d = vaqt_qosh(soat, daqiqa, -15)
            keyin_s, keyin_d = vaqt_qosh(soat, daqiqa, 20)
            try:
                sched.remove_job(f"namoz_oldin_{namoz}")
            except Exception:
                pass
            sched.add_job(namoz_eslatma, "cron", hour=oldin_s, minute=oldin_d,
                          args=[app, namoz.capitalize()], id=f"namoz_oldin_{namoz}", replace_existing=True)
            try:
                sched.remove_job(f"namoz_keyin_{namoz}")
            except Exception:
                pass
            sched.add_job(namoz_oqildimi_sorov, "cron", hour=keyin_s, minute=keyin_d,
                          args=[app, namoz.capitalize()], id=f"namoz_keyin_{namoz}", replace_existing=True)
            if namoz == "bomdod":
                try:
                    sched.remove_job("uyg_onish")
                except Exception:
                    pass
                sched.add_job(uyg_onish_eslatma, "cron", hour=oldin_s, minute=oldin_d,
                              args=[app], id="uyg_onish", replace_existing=True)
            logger.info(f"{namoz} rejalashtirildi: {oldin_s}:{oldin_d:02d}")
        except Exception as e:
            logger.error(f"Namoz xatosi {namoz}: {e}")


# ─── TUGMALAR ─────────────────────────────────────────────────
def asosiy_menyu(user_id=None):
    tugmalar = [
        ["➕ Kirim qo'shish", "➖ Chiqim qo'shish"],
        ["📊 Bugungi hisobot", "📅 Oylik hisobot"],
        ["🗓 Haftalik hisobot", "📂 Kategoriya bo'yicha"],
        ["🕌 Namoz vaqtlari", "💸 Kunlik limit"],
        ["💰 Oylik byudjet", "📒 Qarz ro'yxati"],
        ["➕ Yangi qarz", "📅 Muhim sanalar"],
        ["📥 Excel hisobot", "🗑 Oxirgi yozuvni o'chirish"],
        ["🔔 Eslatmalar", "❓ Yordam"],
    ]
    if user_id == ADMIN_ID:
        tugmalar.append(["👑 Admin panel", "📢 Xabar yuborish"])
    return ReplyKeyboardMarkup(tugmalar, resize_keyboard=True)


def bekor_qilish_menyu():
    return ReplyKeyboardMarkup([["🚫 Bekor qilish"]], resize_keyboard=True)


def kategoriya_menyu():
    qatorlar = [KATEGORIYALAR[i:i+2] for i in range(0, len(KATEGORIYALAR), 2)]
    qatorlar.append(["🚫 Bekor qilish"])
    return ReplyKeyboardMarkup(qatorlar, resize_keyboard=True)


def miqdor_menyu():
    qatorlar = [TAYYOR_MIQDORLAR[i:i+2] for i in range(0, len(TAYYOR_MIQDORLAR), 2)]
    qatorlar.append(["🚫 Bekor qilish"])
    return ReplyKeyboardMarkup(qatorlar, resize_keyboard=True)


# ─── /start ───────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    royxatdan_otkazish(user.id, user.first_name, user.username)
    nom, vaqt = keyingi_namoz()
    if user.id == ADMIN_ID:
        jami, faol, kirim, chiqim, qarzlar = admin_statistika()
        matn = (
            f"👑 Xush kelibsiz, Jahongir aka!\n\n"
            f"📊 Bugungi holat:\n"
            f"👥 Foydalanuvchilar: {jami} ta\n"
            f"🟢 Bugun faol: {faol} ta\n"
            f"❌ Bugungi chiqim: {chiqim:,.0f} som\n"
            f"📒 Ochiq qarzlar: {qarzlar} ta\n\n"
            f"🕌 Keyingi namoz: {nom} — {vaqt}\n\n"
            f"Quyidagi tugmalardan foydalaning 👇"
        )
    else:
        rows = bugungi_hisobot_db(user.id)
        kirim = sum(r[1] for r in rows if r[0] == "kirim")
        chiqim = sum(r[1] for r in rows if r[0] == "chiqim")
        matn = (
            f"👋 Assalomu alaykum, {user.first_name}!\n\n"
            f"✨ Jahongir akadan foydalisi ✨\n\n"
            f"📊 Bugungi hisobingiz:\n"
            f"✅ Kirim: {kirim:,.0f} som\n"
            f"❌ Chiqim: {chiqim:,.0f} som\n"
            f"💰 Qoldiq: {kirim-chiqim:,.0f} som\n\n"
            f"🕌 Keyingi namoz: {nom} — {vaqt}\n\n"
            f"Quyidagi tugmalardan foydalaning 👇"
        )
    await update.message.reply_text(matn, reply_markup=asosiy_menyu(user.id))


# ─── KIRIM ────────────────────────────────────────────────────
async def kirim_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("➕ Miqdorni tanlang yoki kiriting:", reply_markup=miqdor_menyu())
    return KIRIM_MIQDOR


async def kirim_miqdor_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    matn = update.message.text.replace(",", "").replace(" ", "")
    if "yozaman" in matn:
        await update.message.reply_text("💰 Miqdorni kiriting:", reply_markup=bekor_qilish_menyu())
        return KIRIM_MIQDOR
    try:
        miqdor = float(matn)
        if miqdor <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❗ To'g'ri son kiriting:")
        return KIRIM_MIQDOR
    context.user_data["kirim_miqdor"] = miqdor
    await update.message.reply_text("📝 Tavsif kiriting (masalan: maosh):", reply_markup=bekor_qilish_menyu())
    return KIRIM_TAVSIF


async def kirim_tavsif_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    miqdor = context.user_data.pop("kirim_miqdor")
    tavsif = update.message.text.strip() or "Kirim"
    user_id = update.effective_user.id
    qosh_tranzaksiya(user_id, "kirim", miqdor, None, tavsif)
    await update.message.reply_text(
        f"✅ Kirim qo'shildi: {miqdor:,.0f} som — {tavsif}",
        reply_markup=asosiy_menyu(user_id)
    )
    return ConversationHandler.END


# ─── CHIQIM — faqat miqdor + kategoriya (izoh yo'q) ──────────
async def chiqim_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("➖ Miqdorni tanlang yoki kiriting:", reply_markup=miqdor_menyu())
    return CHIQIM_MIQDOR


async def chiqim_miqdor_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    matn = update.message.text.replace(",", "").replace(" ", "")
    if "yozaman" in matn:
        await update.message.reply_text("💰 Miqdorni kiriting:", reply_markup=bekor_qilish_menyu())
        return CHIQIM_MIQDOR
    try:
        miqdor = float(matn)
        if miqdor <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❗ To'g'ri son kiriting:")
        return CHIQIM_MIQDOR
    context.user_data["chiqim_miqdor"] = miqdor
    await update.message.reply_text("📂 Kategoriyani tanlang:", reply_markup=kategoriya_menyu())
    return CHIQIM_KATEGORIYA


async def chiqim_kategoriya_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    if update.message.text not in KATEGORIYALAR:
        await update.message.reply_text("❗ Ro'yxatdan tanlang:", reply_markup=kategoriya_menyu())
        return CHIQIM_KATEGORIYA
    miqdor = context.user_data.pop("chiqim_miqdor")
    kategoriya = update.message.text
    user_id = update.effective_user.id
    qosh_tranzaksiya(user_id, "chiqim", miqdor, kategoriya, kategoriya)
    await update.message.reply_text(
        f"❌ Chiqim: {miqdor:,.0f} som — {kategoriya}",
        reply_markup=asosiy_menyu(user_id)
    )
    limit_q = limit_olish(user_id)
    if limit_q > 0:
        jami = bugungi_chiqim_jami(user_id)
        if jami >= limit_q and not limit_ogohlantirish_yuborilganmi(user_id, date.today()):
            limit_ogohlantirish_belgilash(user_id, date.today())
            await update.message.reply_text(
                f"⚠️ Kunlik limit oshib ketdi!\nLimit: {limit_q:,.0f} som\nBugungi chiqim: {jami:,.0f} som"
            )
    return ConversationHandler.END


async def bekor_qilish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("🚫 Bekor qilindi.", reply_markup=asosiy_menyu(update.effective_user.id))
    return ConversationHandler.END


# ─── LIMIT ────────────────────────────────────────────────────
async def limit_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    limit_q = limit_olish(update.effective_user.id)
    matn = "💸 Kunlik limit kiriting (som):\n"
    if limit_q > 0:
        matn += f"Joriy: {limit_q:,.0f} som\n"
    matn += "0 — limit o'chirish"
    await update.message.reply_text(matn, reply_markup=bekor_qilish_menyu())
    return LIMIT_KIRITISH


async def limit_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    try:
        limit_q = float(update.message.text.replace(" ", "").replace(",", ""))
        if limit_q < 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❗ To'g'ri son kiriting:")
        return LIMIT_KIRITISH
    user_id = update.effective_user.id
    limit_saqlash(user_id, limit_q)
    matn = "✅ Limit o'chirildi." if limit_q == 0 else f"✅ Limit: {limit_q:,.0f} som"
    await update.message.reply_text(matn, reply_markup=asosiy_menyu(user_id))
    return ConversationHandler.END


# ─── BYUDJET ──────────────────────────────────────────────────
async def byudjet_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    byudjet = byudjet_olish(update.effective_user.id)
    matn = "💰 Oylik byudjet kiriting (som):\n"
    if byudjet > 0:
        matn += f"Joriy: {byudjet:,.0f} som\n"
    matn += "0 — byudjet o'chirish"
    await update.message.reply_text(matn, reply_markup=bekor_qilish_menyu())
    return BYUDJET_KIRITISH


async def byudjet_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    try:
        byudjet = float(update.message.text.replace(" ", "").replace(",", ""))
        if byudjet < 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❗ To'g'ri son kiriting:")
        return BYUDJET_KIRITISH
    user_id = update.effective_user.id
    byudjet_saqlash(user_id, byudjet)
    matn = "✅ Byudjet o'chirildi." if byudjet == 0 else f"✅ Byudjet: {byudjet:,.0f} som"
    await update.message.reply_text(matn, reply_markup=asosiy_menyu(user_id))
    return ConversationHandler.END


# ─── QARZ ─────────────────────────────────────────────────────
async def qarz_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("📤 Men berdim", callback_data="qarz_tur_berdi"),
        InlineKeyboardButton("📥 Men oldim", callback_data="qarz_tur_oldi")
    ]])
    await update.message.reply_text("📒 Yangi qarz — tur tanlang:", reply_markup=keyboard)
    return QARZ_TUR


async def qarz_tur_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tur = "berdi" if query.data == "qarz_tur_berdi" else "oldi"
    context.user_data["qarz_tur"] = tur
    await query.edit_message_text("📤 Kimga berdingiz? Ism:" if tur == "berdi" else "📥 Kimdan oldingiz? Ism:")
    return QARZ_ISM


async def qarz_ism_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    context.user_data["qarz_ism"] = update.message.text.strip()
    await update.message.reply_text("💰 Miqdor (som):", reply_markup=bekor_qilish_menyu())
    return QARZ_MIQDOR


async def qarz_miqdor_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    try:
        miqdor = float(update.message.text.replace(" ", "").replace(",", ""))
        if miqdor <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❗ To'g'ri son kiriting:")
        return QARZ_MIQDOR
    context.user_data["qarz_miqdor"] = miqdor
    await update.message.reply_text("📅 Qaytarish sanasi (25.07.2026):", reply_markup=bekor_qilish_menyu())
    return QARZ_ESLATMA


async def qarz_sana_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    try:
        sana = datetime.strptime(update.message.text.strip(), "%d.%m.%Y").date()
        if sana < date.today():
            await update.message.reply_text("❗ Kelajakdagi sana kiriting:")
            return QARZ_ESLATMA
    except ValueError:
        await update.message.reply_text("❗ Format xato. Masalan: 25.07.2026")
        return QARZ_ESLATMA
    context.user_data["qarz_sana"] = str(sana)
    await update.message.reply_text("📝 Izoh (ixtiyoriy, yo'q — tire):", reply_markup=bekor_qilish_menyu())
    return QARZ_IZOH


async def qarz_izoh_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    user_id = update.effective_user.id
    izoh = update.message.text.strip()
    if izoh in ["-", ".", "yo'q", "yoq"]:
        izoh = ""
    ism = context.user_data.pop("qarz_ism")
    miqdor = context.user_data.pop("qarz_miqdor")
    tur = context.user_data.pop("qarz_tur")
    sana = context.user_data.pop("qarz_sana")
    qarz_qosh_db(user_id, ism, miqdor, tur, sana, izoh)
    tur_matn = f"{ism} ga berdingiz" if tur == "berdi" else f"{ism} dan oldingiz"
    await update.message.reply_text(
        f"✅ Saqlandi!\n{tur_matn}\n{miqdor:,.0f} som | {sana}",
        reply_markup=asosiy_menyu(user_id)
    )
    return ConversationHandler.END


# ─── MUHIM SANALAR ────────────────────────────────────────────
async def sana_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sanalar = sanalar_royxat_db(user_id)
    matn = "📅 Muhim sanalar\n\n"
    if sanalar:
        for sid, ism, sana, tavsif in sanalar:
            matn += f"[{sid}] {ism} — {sana}\n"
    else:
        matn += "Hali yo'q.\n"
    matn += "\nYangi qo'shish uchun ism kiriting:"
    await update.message.reply_text(matn, reply_markup=bekor_qilish_menyu())
    return SANA_ISM


async def sana_ism_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    context.user_data["sana_ism"] = update.message.text.strip()
    await update.message.reply_text("📅 Sana (15.03.2026):", reply_markup=bekor_qilish_menyu())
    return SANA_SANA


async def sana_sana_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    try:
        sana = datetime.strptime(update.message.text.strip(), "%d.%m.%Y").date()
    except ValueError:
        await update.message.reply_text("❗ Format xato. Masalan: 15.03.2026")
        return SANA_SANA
    context.user_data["sana_sana"] = str(sana)
    await update.message.reply_text("📝 Izoh (ixtiyoriy, yo'q — tire):", reply_markup=bekor_qilish_menyu())
    return SANA_TAVSIF


async def sana_tavsif_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    user_id = update.effective_user.id
    tavsif = update.message.text.strip()
    if tavsif in ["-", ".", "yo'q"]:
        tavsif = ""
    ism = context.user_data.pop("sana_ism")
    sana = context.user_data.pop("sana_sana")
    sana_qosh_db(user_id, ism, sana, tavsif)
    await update.message.reply_text(f"✅ Saqlandi! {ism} — {sana}", reply_markup=asosiy_menyu(user_id))
    return ConversationHandler.END


# ─── ADMIN XABAR ──────────────────────────────────────────────
async def xabar_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("Ruxsat yo'q.")
        return ConversationHandler.END
    await update.message.reply_text("📢 Xabar yozing:", reply_markup=bekor_qilish_menyu())
    return XABAR_MATN


async def xabar_matn_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🚫 Bekor qilish":
        return await bekor_qilish(update, context)
    matn = update.message.text.strip()
    yuborildi = 0
    for user_id, ism, username, limit_q, eslatma in barcha_userlar():
        try:
            await context.bot.send_message(chat_id=user_id, text=f"📢 Admin xabari:\n\n{matn}")
            yuborildi += 1
        except Exception:
            pass
    await update.message.reply_text(f"✅ {yuborildi} ta foydalanuvchiga yuborildi.", reply_markup=asosiy_menyu(ADMIN_ID))
    return ConversationHandler.END


# ─── HISOBOTLAR ───────────────────────────────────────────────
async def bugun_hisobot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    rows = bugungi_hisobot_db(user_id)
    if not rows:
        await update.message.reply_text("📭 Bugun hech narsa kiritilmagan.")
        return
    kirim = sum(r[1] for r in rows if r[0] == "kirim")
    chiqim = sum(r[1] for r in rows if r[0] == "chiqim")
    limit_q = limit_olish(user_id)
    matn = f"📊 Bugun — {date.today().strftime('%d.%m.%Y')}\n"
    matn += "─────────────────\n"
    matn += f"✅ Kirim:  {kirim:,.0f} som\n"
    matn += f"❌ Chiqim: {chiqim:,.0f} som\n"
    matn += "─────────────────\n"
    matn += f"💰 Qoldiq: {kirim-chiqim:,.0f} som\n"
    if limit_q > 0:
        foiz = min(100, round(chiqim / limit_q * 100))
        matn += f"🎯 Limit: {foiz}% ({chiqim:,.0f}/{limit_q:,.0f})\n"
    await update.message.reply_text(matn)


async def oy_hisobot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    oy = datetime.now().strftime("%Y-%m")
    rows = oylik_hisobot_db(user_id, oy)
    if not rows:
        await update.message.reply_text("📭 Bu oy hech narsa kiritilmagan.")
        return
    kirim = next((r[1] for r in rows if r[0] == "kirim"), 0)
    chiqim = next((r[1] for r in rows if r[0] == "chiqim"), 0)
    birinchi = date.today().replace(day=1)
    otgan_oy = (birinchi - timedelta(days=1)).strftime("%Y-%m")
    otgan_rows = oylik_hisobot_db(user_id, otgan_oy)
    otgan_chiqim = next((r[1] for r in otgan_rows if r[0] == "chiqim"), 0)
    byudjet = byudjet_olish(user_id)
    matn = f"📅 {datetime.now().strftime('%B %Y')}\n"
    matn += "─────────────────\n"
    matn += f"✅ Kirim:  {kirim:,.0f} som\n"
    matn += f"❌ Chiqim: {chiqim:,.0f} som\n"
    matn += "─────────────────\n"
    matn += f"💰 Qoldiq: {kirim-chiqim:,.0f} som\n"
    if byudjet > 0:
        foiz = min(100, round(chiqim / byudjet * 100))
        matn += f"💰 Byudjet: {foiz}%\n"
    if otgan_chiqim > 0:
        farq = chiqim - otgan_chiqim
        belgi = "📈" if farq > 0 else "📉"
        kyk = "ko'p" if farq > 0 else "kam"
        matn += f"{belgi} O'tgan oyga: {abs(farq):,.0f} som {kyk}\n"
    await update.message.reply_text(matn)


async def hafta_hisobot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    bir_hafta = date.today() - timedelta(days=7)
    rows = davr_hisobot_db(user_id, bir_hafta)
    if not rows:
        await update.message.reply_text("📭 Bu hafta hech narsa kiritilmagan.")
        return
    kirim = sum(r[1] for r in rows if r[0] == "kirim")
    chiqim = sum(r[1] for r in rows if r[0] == "chiqim")
    matn = f"🗓 {bir_hafta.strftime('%d.%m')} — {date.today().strftime('%d.%m')}\n"
    matn += "─────────────────\n"
    matn += f"✅ Kirim:  {kirim:,.0f} som\n"
    matn += f"❌ Chiqim: {chiqim:,.0f} som\n"
    matn += "─────────────────\n"
    matn += f"💰 Qoldiq: {kirim-chiqim:,.0f} som\n"
    await update.message.reply_text(matn)


async def kategoriya_hisobot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    rows = kategoriya_hisobot_db(user_id)
    if not rows:
        await update.message.reply_text("📭 Bu oy chiqim yo'q.")
        return
    matn = f"📂 {datetime.now().strftime('%B %Y')}\n\n"
    jami = sum(r[1] for r in rows)
    for kategoriya, miqdor in rows:
        foiz = round(miqdor / jami * 100)
        matn += f"{kategoriya or 'Boshqa'}: {miqdor:,.0f} som ({foiz}%)\n"
    matn += f"\nJami: {jami:,.0f} som"
    await update.message.reply_text(matn)


async def namoz_vaqtlari_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    vaqtlar = namoz_vaqtlarini_ol()
    if not vaqtlar:
        await update.message.reply_text("❗ Namoz vaqtlarini olishda xatolik.")
        return
    emoji = {"bomdod": "🌙", "peshin": "☀️", "asr": "🌤", "shom": "🌇", "xufton": "🌃"}
    matn = f"🕌 Buxoro — {date.today().strftime('%d.%m.%Y')}\n\n"
    for nom, vaqt in vaqtlar.items():
        matn += f"{emoji.get(nom, '')} {nom.capitalize()}: {vaqt}\n"
    await update.message.reply_text(matn)


async def excel_hisobot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    boshlanish = date.today().replace(day=1)
    rows = davr_hisobot_db(user_id, boshlanish)
    if not rows:
        await update.message.reply_text("📭 Bu oy malumot yo'q.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    sarlavha_font = Font(bold=True, color="FFFFFF")
    sarlavha_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    ws.append(["Sana", "Vaqt", "Turi", "Kategoriya", "Tavsif", "Miqdor"])
    for h in ws[1]:
        h.font = sarlavha_font
        h.fill = sarlavha_fill
        h.alignment = Alignment(horizontal="center")
    jami_kirim = jami_chiqim = 0
    for tur, miqdor, kategoriya, tavsif, sana, vaqt in rows:
        ws.append([sana, vaqt, "Kirim" if tur == "kirim" else "Chiqim", kategoriya or "-", tavsif, miqdor])
        if tur == "kirim":
            jami_kirim += miqdor
        else:
            jami_chiqim += miqdor
    ws.append([])
    ws.append(["", "", "", "", "Jami kirim:", jami_kirim])
    ws.append(["", "", "", "", "Jami chiqim:", jami_chiqim])
    ws.append(["", "", "", "", "Qoldiq:", jami_kirim - jami_chiqim])
    for ustun, kenglik in zip("ABCDEF", [12, 8, 10, 16, 28, 16]):
        ws.column_dimensions[ustun].width = kenglik
    fayl = os.path.join(BASE_DIR, f"h_{user_id}.xlsx")
    wb.save(fayl)
    oy = datetime.now().strftime("%B %Y")
    with open(fayl, "rb") as f:
        await update.message.reply_document(document=f, filename=f"hisobot_{oy}.xlsx",
                                             caption=f"📥 {oy} oyi hisoboti")
    try:
        os.remove(fayl)
    except Exception:
        pass


async def qarz_royxat_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    qarzlar = qarz_royxat_db(user_id)
    if not qarzlar:
        await update.message.reply_text("📭 Ochiq qarzlar yo'q.\nYangi qarz qo'shish — Yangi qarz tugmasi.")
        return
    berganlar = [q for q in qarzlar if q[3] == "berdi"]
    olganlar = [q for q in qarzlar if q[3] == "oldi"]
    matn = "📒 Qarz daftari\n\n"
    if berganlar:
        jami = sum(q[2] for q in berganlar)
        matn += f"📤 Men berganlarim — {jami:,.0f} som\n"
        for qid, ism, miqdor, tur, sana, izoh in berganlar:
            belgi = "⚠️" if sana <= str(date.today()) else "🟢"
            matn += f"{belgi} [{qid}] {ism} — {miqdor:,.0f} | {sana}\n"
        matn += "\n"
    if olganlar:
        jami = sum(q[2] for q in olganlar)
        matn += f"📥 Men olganlarim — {jami:,.0f} som\n"
        for qid, ism, miqdor, tur, sana, izoh in olganlar:
            belgi = "⚠️" if sana <= str(date.today()) else "🟢"
            matn += f"{belgi} [{qid}] {ism} — {miqdor:,.0f} | {sana}\n"
    matn += "\nQaytarildi: /qaytarildi 3"
    await update.message.reply_text(matn)


async def qarz_qaytarildi_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    try:
        qarz_id = int(context.args[0])
    except (IndexError, ValueError):
        await update.message.reply_text("Yozing: /qaytarildi 3")
        return
    row = qarz_yop(qarz_id, user_id)
    if row:
        ism, miqdor, tur = row
        tur_matn = f"{ism} ga bergan" if tur == "berdi" else f"{ism} dan olgan"
        await update.message.reply_text(
            f"✅ Qarz yopildi!\n{tur_matn}\n{miqdor:,.0f} som\n{date.today().strftime('%d.%m.%Y')}"
        )
    else:
        await update.message.reply_text("Topilmadi.")


async def oxirgi_ochirish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    row = oxirgi_yozuvni_ochirish(update.effective_user.id)
    if not row:
        await update.message.reply_text("📭 O'chirish uchun yozuv yo'q.")
        return
    tur_matn = "Kirim" if row[1] == "kirim" else "Chiqim"
    await update.message.reply_text(f"🗑 O'chirildi: {tur_matn} — {row[2]:,.0f} som ({row[3]})")


async def eslatmalar_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    holat = eslatma_holati(user_id)
    matn = "🔔 Eslatmalar: " + ("✅ Yoqilgan" if holat == 1 else "🔇 O'chirilgan")
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("🔇 O'chirish" if holat == 1 else "🔔 Yoqish", callback_data="eslatma_toggle")
    ]])
    await update.message.reply_text(matn, reply_markup=keyboard)


async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("Ruxsat yo'q.")
        return
    jami, faol, kirim, chiqim, qarzlar = admin_statistika()
    matn = "👑 Admin Panel\n-----------------\n"
    matn += f"Jami: {jami} ta\nBugun faol: {faol} ta\n"
    matn += f"Oylik kirim: {kirim:,.0f} som\nOylik chiqim: {chiqim:,.0f} som\n"
    matn += f"Qoldiq: {kirim-chiqim:,.0f} som\nOchiq qarzlar: {qarzlar} ta\n"
    matn += "-----------------\nFoydalanuvchilar:\n"
    for uid, ism, username, limit_q, eslatma in barcha_userlar():
        uname = f"@{username}" if username else "-"
        e = "ON" if eslatma == 1 else "OFF"
        matn += f"- {str(ism).replace('_', ' ')} ({uname}) [{e}]\n"
    if len(matn) > 4000:
        for i in range(0, len(matn), 4000):
            await update.message.reply_text(matn[i:i+4000])
    else:
        await update.message.reply_text(matn)


async def yordam(update: Update, context: ContextTypes.DEFAULT_TYPE):
    matn = (
        "📋 Barcha imkoniyatlar:\n\n"
        "➕ Kirim / ➖ Chiqim — kiritish\n"
        "📊 Bugungi / 📅 Oylik / 🗓 Haftalik hisobot\n"
        "📂 Kategoriya boyicha tahlil\n"
        "🕌 Namoz vaqtlari (Buxoro)\n"
        "💸 Kunlik limit / 💰 Oylik byudjet\n"
        "📒 Qarz daftari\n"
        "📅 Muhim sanalar\n"
        "📥 Excel hisobot\n"
        "🔔 Eslatmalarni yoqish/ochirish\n\n"
        "Avtomatik:\n"
        "07:00 motivatsion xabar\n"
        "08:00 muhim sanalar\n"
        "09:00 qarz eslatma\n"
        "12:00 soglik maslahati\n"
        "Har namozdan 15 daqiqa oldin\n"
        "22:30 uyqu eslatmasi\n"
        "00:00 kunlik hisobot\n"
    )
    await update.message.reply_text(matn)


# ─── CALLBACKS ────────────────────────────────────────────────
async def eslatma_toggle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    yangi = eslatma_almashtirish(query.from_user.id)
    matn = "🔔 Yoqildi" if yangi == 1 else "🔇 O'chirildi"
    await query.answer(matn)
    await query.edit_message_text(matn)


async def namoz_javob_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data.startswith("namoz_ha_"):
        namoz = query.data.replace("namoz_ha_", "")
        await query.edit_message_text(f"✅ {namoz} namozi oqildi. Alloh qabul qilsin!")
    elif query.data.startswith("namoz_yoq_"):
        namoz = query.data.replace("namoz_yoq_", "")
        await query.edit_message_text(f"Imkon bolsa oqing — vaqt bor hali!")


# ─── MATN ROUTER ──────────────────────────────────────────────
async def matn_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    matn = update.message.text.strip()
    mapping = {
        "📊 Bugungi hisobot": bugun_hisobot,
        "📅 Oylik hisobot": oy_hisobot,
        "🗓 Haftalik hisobot": hafta_hisobot,
        "📂 Kategoriya bo'yicha": kategoriya_hisobot,
        "🕌 Namoz vaqtlari": namoz_vaqtlari_cmd,
        "📥 Excel hisobot": excel_hisobot,
        "🗑 Oxirgi yozuvni o'chirish": oxirgi_ochirish,
        "🔔 Eslatmalar": eslatmalar_cmd,
        "❓ Yordam": yordam,
        "👑 Admin panel": admin_panel,
        "📒 Qarz ro'yxati": qarz_royxat_cmd,
    }
    funksiya = mapping.get(matn)
    if funksiya:
        await funksiya(update, context)


# ─── MAIN ─────────────────────────────────────────────────────
def main():
    global scheduler, application
    init_db()
    application = Application.builder().token(BOT_TOKEN).build()

    kirim_conv = ConversationHandler(
        entry_points=[CommandHandler("kirim", kirim_boshlash),
                      MessageHandler(filters.Regex("^➕ Kirim qo'shish$"), kirim_boshlash)],
        states={
            KIRIM_MIQDOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, kirim_miqdor_qabul)],
            KIRIM_TAVSIF: [MessageHandler(filters.TEXT & ~filters.COMMAND, kirim_tavsif_qabul)],
        },
        fallbacks=[MessageHandler(filters.Regex("^🚫 Bekor qilish$"), bekor_qilish)],
    )

    chiqim_conv = ConversationHandler(
        entry_points=[CommandHandler("chiqim", chiqim_boshlash),
                      MessageHandler(filters.Regex("^➖ Chiqim qo'shish$"), chiqim_boshlash)],
        states={
            CHIQIM_MIQDOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, chiqim_miqdor_qabul)],
            CHIQIM_KATEGORIYA: [MessageHandler(filters.TEXT & ~filters.COMMAND, chiqim_kategoriya_qabul)],
        },
        fallbacks=[MessageHandler(filters.Regex("^🚫 Bekor qilish$"), bekor_qilish)],
    )

    limit_conv = ConversationHandler(
        entry_points=[CommandHandler("limit", limit_boshlash),
                      MessageHandler(filters.Regex("^💸 Kunlik limit$"), limit_boshlash)],
        states={LIMIT_KIRITISH: [MessageHandler(filters.TEXT & ~filters.COMMAND, limit_qabul)]},
        fallbacks=[MessageHandler(filters.Regex("^🚫 Bekor qilish$"), bekor_qilish)],
    )

    byudjet_conv = ConversationHandler(
        entry_points=[CommandHandler("byudjet", byudjet_boshlash),
                      MessageHandler(filters.Regex("^💰 Oylik byudjet$"), byudjet_boshlash)],
        states={BYUDJET_KIRITISH: [MessageHandler(filters.TEXT & ~filters.COMMAND, byudjet_qabul)]},
        fallbacks=[MessageHandler(filters.Regex("^🚫 Bekor qilish$"), bekor_qilish)],
    )

    qarz_conv = ConversationHandler(
        entry_points=[CommandHandler("qarz", qarz_boshlash),
                      MessageHandler(filters.Regex("^➕ Yangi qarz$"), qarz_boshlash)],
        states={
            QARZ_TUR: [CallbackQueryHandler(qarz_tur_callback, pattern="^qarz_tur_")],
            QARZ_ISM: [MessageHandler(filters.TEXT & ~filters.COMMAND, qarz_ism_qabul)],
            QARZ_MIQDOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, qarz_miqdor_qabul)],
            QARZ_ESLATMA: [MessageHandler(filters.TEXT & ~filters.COMMAND, qarz_sana_qabul)],
            QARZ_IZOH: [MessageHandler(filters.TEXT & ~filters.COMMAND, qarz_izoh_qabul)],
        },
        fallbacks=[MessageHandler(filters.Regex("^🚫 Bekor qilish$"), bekor_qilish)],
    )

    sana_conv = ConversationHandler(
        entry_points=[CommandHandler("sana", sana_boshlash),
                      MessageHandler(filters.Regex("^📅 Muhim sanalar$"), sana_boshlash)],
        states={
            SANA_ISM: [MessageHandler(filters.TEXT & ~filters.COMMAND, sana_ism_qabul)],
            SANA_SANA: [MessageHandler(filters.TEXT & ~filters.COMMAND, sana_sana_qabul)],
            SANA_TAVSIF: [MessageHandler(filters.TEXT & ~filters.COMMAND, sana_tavsif_qabul)],
        },
        fallbacks=[MessageHandler(filters.Regex("^🚫 Bekor qilish$"), bekor_qilish)],
    )

    xabar_conv = ConversationHandler(
        entry_points=[CommandHandler("xabar", xabar_boshlash),
                      MessageHandler(filters.Regex("^📢 Xabar yuborish$"), xabar_boshlash)],
        states={XABAR_MATN: [MessageHandler(filters.TEXT & ~filters.COMMAND, xabar_matn_qabul)]},
        fallbacks=[MessageHandler(filters.Regex("^🚫 Bekor qilish$"), bekor_qilish)],
    )

    application.add_handler(CommandHandler("start", start))
    application.add_handler(kirim_conv)
    application.add_handler(chiqim_conv)
    application.add_handler(limit_conv)
    application.add_handler(byudjet_conv)
    application.add_handler(qarz_conv)
    application.add_handler(sana_conv)
    application.add_handler(xabar_conv)
    application.add_handler(CommandHandler("bugun", bugun_hisobot))
    application.add_handler(CommandHandler("oy", oy_hisobot))
    application.add_handler(CommandHandler("hafta", hafta_hisobot))
    application.add_handler(CommandHandler("kategoriya", kategoriya_hisobot))
    application.add_handler(CommandHandler("namoz", namoz_vaqtlari_cmd))
    application.add_handler(CommandHandler("excel", excel_hisobot))
    application.add_handler(CommandHandler("qarzlar", qarz_royxat_cmd))
    application.add_handler(CommandHandler("qaytarildi", qarz_qaytarildi_cmd))
    application.add_handler(CommandHandler("admin", admin_panel))
    application.add_handler(CommandHandler("yordam", yordam))
    application.add_handler(CallbackQueryHandler(eslatma_toggle_callback, pattern="^eslatma_toggle$"))
    application.add_handler(CallbackQueryHandler(namoz_javob_callback, pattern="^namoz_(ha|yoq)_"))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, matn_router))

    scheduler = AsyncIOScheduler(timezone="Asia/Tashkent")
    rejalashtirish(scheduler, application)
    namoz_rejalashtir(scheduler, application)
    scheduler.start()

    logger.info("Bot ishga tushdi!")

    async def health_check(request):
        # Ping kelganda schedulerni tekshir, to'xtagan bo'lsa qayta ishga tushir
        global scheduler, application
        if not scheduler.running:
            logger.warning("Scheduler to'xtagan — qayta ishga tushirilmoqda!")
            scheduler = AsyncIOScheduler(timezone="Asia/Tashkent")
            rejalashtirish(scheduler, application)
            namoz_rejalashtir(scheduler, application)
            scheduler.start()
        return web.Response(text="OK")

    async def run_web():
        app_web = web.Application()
        app_web.router.add_get("/", health_check)
        runner = web.AppRunner(app_web)
        await runner.setup()
        site = web.TCPSite(runner, "0.0.0.0", int(os.environ.get("PORT", 8080)))
        await site.start()
        logger.info("Web server ishga tushdi!")

    async def run_bot():
        await application.initialize()
        await application.start()
        await application.updater.start_polling(drop_pending_updates=True)
        logger.info("Bot polling boshlandi!")

    async def main_async():
        await asyncio.gather(run_web(), run_bot())
        await asyncio.Event().wait()

    asyncio.run(main_async())


if __name__ == "__main__":
    main()
